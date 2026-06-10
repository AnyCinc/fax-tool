"""複数Excelリストの結合・分割ページ"""
import os
import sys
import zipfile
import csv
from io import BytesIO, StringIO
import streamlit as st
import openpyxl
from openpyxl.utils import get_column_letter
from copy import copy

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

st.title("📋 リスト結合・分割")

tab_merge, tab_split = st.tabs(["🔗 リスト結合", "✂️ リスト分割"])

# ===================== リスト結合タブ =====================
with tab_merge:
    st.caption("複数のExcelファイルを1つに結合します（列はそのまま維持）")

    uploaded_files = st.file_uploader(
        "Excelファイルを複数選択",
        type=["xlsx", "xls"],
        accept_multiple_files=True,
    )

    if uploaded_files:
        st.info(f"📁 {len(uploaded_files)}件のファイルを選択中")

        header_row = st.number_input("ヘッダー行（1行目がヘッダーなら1）", min_value=0, max_value=10, value=1)
        skip_header = st.checkbox("2つ目以降のファイルのヘッダーをスキップする", value=True)

        if st.button("📊 プレビュー", use_container_width=True):
            total_rows = 0
            for i, f in enumerate(uploaded_files):
                wb = openpyxl.load_workbook(BytesIO(f.read()), data_only=True)
                f.seek(0)
                ws = wb.active

                last_row = 0
                for row in range(ws.max_row, 0, -1):
                    if any(ws.cell(row, c).value for c in range(1, ws.max_column + 1)):
                        last_row = row
                        break

                data_start = (header_row + 1) if header_row > 0 else 1
                data_rows = last_row - data_start + 1 if last_row >= data_start else 0

                st.markdown(f"**{i+1}. {f.name}** — {data_rows}行 × {ws.max_column}列")
                total_rows += data_rows

                if header_row > 0 and i == 0:
                    headers = [ws.cell(header_row, c).value or "" for c in range(1, ws.max_column + 1)]
                    st.write("ヘッダー:", headers)

            st.success(f"**合計: {total_rows}行** のデータが結合されます")

        st.divider()

        output_name = st.text_input("出力ファイル名", value="結合リスト.xlsx")

        if st.button("🔗 結合して生成", type="primary", use_container_width=True):
            with st.spinner("結合中..."):
                try:
                    result_wb = openpyxl.Workbook()
                    result_ws = result_wb.active
                    result_ws.title = "結合データ"
                    current_row = 1

                    for i, f in enumerate(uploaded_files):
                        wb = openpyxl.load_workbook(BytesIO(f.read()), data_only=True)
                        f.seek(0)
                        ws = wb.active

                        last_row = 0
                        for row in range(ws.max_row, 0, -1):
                            if any(ws.cell(row, c).value for c in range(1, ws.max_column + 1)):
                                last_row = row
                                break

                        if i == 0:
                            start_row = 1
                        else:
                            if skip_header and header_row > 0:
                                start_row = header_row + 1
                            else:
                                start_row = 1

                        for src_row in range(start_row, last_row + 1):
                            for col in range(1, ws.max_column + 1):
                                src_cell = ws.cell(src_row, col)
                                dest_cell = result_ws.cell(current_row, col)
                                dest_cell.value = src_cell.value

                                if src_cell.has_style:
                                    dest_cell.font = copy(src_cell.font)
                                    dest_cell.fill = copy(src_cell.fill)
                                    dest_cell.border = copy(src_cell.border)
                                    dest_cell.alignment = copy(src_cell.alignment)
                                    dest_cell.number_format = src_cell.number_format

                            current_row += 1

                        if i == 0:
                            for col in range(1, ws.max_column + 1):
                                letter = get_column_letter(col)
                                if ws.column_dimensions[letter].width:
                                    result_ws.column_dimensions[letter].width = ws.column_dimensions[letter].width

                    output_buffer = BytesIO()
                    result_wb.save(output_buffer)
                    output_buffer.seek(0)

                    total = current_row - 1
                    st.success(f"結合完了！ {total}行のデータ")

                    st.download_button(
                        label="📥 結合Excelをダウンロード",
                        data=output_buffer.getvalue(),
                        file_name=output_name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )

                except Exception as e:
                    st.error(f"結合に失敗しました: {e}")
    else:
        st.markdown("""
        ### 使い方
        1. 上のエリアにExcelファイルを**複数**ドラッグ＆ドロップ
        2. 「プレビュー」で内容を確認
        3. 「結合して生成」で1つのファイルに結合

        > 💡 1つ目のファイルのヘッダーが使われ、2つ目以降のデータが下に追加されます
        """)

# ===================== リスト分割タブ =====================
with tab_split:
    st.caption("CSVファイルを指定件数ごとに分割します")

    split_file = st.file_uploader("CSVファイルを選択", type=["csv"], key="split_csv")

    chunk_size = st.number_input("分割件数（1ファイルあたりの行数）", min_value=100, max_value=100000, value=2500, step=100)

    if split_file:
        # エンコーディング検出して読み込み
        raw = split_file.read()
        split_file.seek(0)

        # よくある日本語エンコーディングを試行
        content = None
        used_encoding = None
        for enc in ["utf-8-sig", "utf-8", "cp932", "shift_jis"]:
            try:
                content = raw.decode(enc)
                used_encoding = enc
                break
            except (UnicodeDecodeError, LookupError):
                continue

        if content is None:
            st.error("ファイルのエンコーディングを判定できませんでした")
        else:
            reader = csv.reader(StringIO(content))
            all_rows = list(reader)

            if len(all_rows) == 0:
                st.warning("CSVにデータがありません")
            else:
                header = all_rows[0]
                data_rows = all_rows[1:]
                total_data = len(data_rows)
                num_files = (total_data + chunk_size - 1) // chunk_size

                st.info(f"📁 **{split_file.name}** — ヘッダー1行 + データ{total_data}行 → **{num_files}ファイル**に分割")
                st.write("ヘッダー:", header[:10], "..." if len(header) > 10 else "")

                # 出力ベース名
                base_name = os.path.splitext(split_file.name)[0]

                if st.button("✂️ 分割して生成", type="primary", use_container_width=True):
                    with st.spinner("分割中..."):
                        zip_buffer = BytesIO()
                        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
                            for idx in range(num_files):
                                start = idx * chunk_size
                                end = min(start + chunk_size, total_data)
                                chunk = data_rows[start:end]

                                # CSVとして書き出し
                                csv_buffer = StringIO()
                                writer = csv.writer(csv_buffer)
                                writer.writerow(header)
                                writer.writerows(chunk)

                                file_name = f"{base_name}_{idx + 1}.csv"
                                zf.writestr(file_name, csv_buffer.getvalue().encode("utf-8-sig"))

                                st.markdown(f"  ✅ **{file_name}** — {len(chunk)}行")

                        zip_buffer.seek(0)
                        st.success(f"分割完了！ {num_files}ファイル生成")

                        st.download_button(
                            label="📥 分割CSVをダウンロード（ZIP）",
                            data=zip_buffer.getvalue(),
                            file_name=f"{base_name}_分割.zip",
                            mime="application/zip",
                            use_container_width=True,
                        )
    else:
        st.markdown("""
        ### 使い方
        1. 上のエリアにCSVファイルをドラッグ＆ドロップ
        2. 分割件数を設定（デフォルト: 2,500件）
        3. 「分割して生成」でZIPファイルとしてダウンロード

        > 💡 各ファイルにヘッダー行が自動で付きます
        """)
