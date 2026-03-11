"""履歴書Excelの読み書きモジュール"""
import os
import shutil
from datetime import datetime
from dataclasses import dataclass, field
from openpyxl import load_workbook

# 履歴書のセルマッピング
CELL_MAP = {
    "furigana": "B3",
    "name": "B4",
    "gender": "E3",
    "birthday": "C6",
    "address": "B7",
    "nationality": "G8",
    "reg_number": "G10",
}

EDUCATION_START_ROW = 13
EDUCATION_HEADER_ROW = 12
WORK_HEADER_ROW = 21
WORK_START_ROW = 22


@dataclass
class EducationEntry:
    year: int = None
    month: int = None
    school: str = ""
    status: str = ""  # 入学/卒業


@dataclass
class WorkEntry:
    year: int = None
    month: int = None
    company: str = ""
    status: str = ""  # 入社/退社
    detail: str = ""  # 業務内容


@dataclass
class ResumeData:
    furigana: str = ""
    name: str = ""
    gender: str = ""
    birthday: datetime = None
    address: str = ""
    nationality: str = ""
    reg_number: str = ""
    education: list = field(default_factory=list)
    work: list = field(default_factory=list)
    source_path: str = ""


def read_resume(filepath: str) -> ResumeData:
    """履歴書Excelを読み込んでResumeDataを返す"""
    wb = load_workbook(filepath)
    ws = wb.active

    data = ResumeData(source_path=filepath)
    data.furigana = ws[CELL_MAP["furigana"]].value or ""
    data.name = ws[CELL_MAP["name"]].value or ""
    data.gender = ws[CELL_MAP["gender"]].value or ""
    data.birthday = ws[CELL_MAP["birthday"]].value
    data.address = ws[CELL_MAP["address"]].value or ""
    data.nationality = ws[CELL_MAP["nationality"]].value or ""
    data.reg_number = ws[CELL_MAP["reg_number"]].value or ""

    # 学歴読み込み (Row 13〜20)
    for row in range(EDUCATION_START_ROW, WORK_HEADER_ROW):
        year = ws.cell(row=row, column=1).value
        month = ws.cell(row=row, column=2).value
        school = ws.cell(row=row, column=3).value
        status = ws.cell(row=row, column=8).value  # H列
        if school:
            # 年月なし＝前の学歴の補足行（学科名など）→スキップ
            if year is None and month is None:
                continue
            data.education.append(EducationEntry(
                year=year, month=month, school=school, status=status or ""
            ))

    # 職歴読み込み (Row 22〜)
    for row in range(WORK_START_ROW, ws.max_row + 1):
        year = ws.cell(row=row, column=1).value
        month = ws.cell(row=row, column=2).value
        company = ws.cell(row=row, column=3).value
        status = ws.cell(row=row, column=8).value
        if company:
            detail = ""
            if company.startswith("業務内容"):
                # This is a detail row for previous entry
                if data.work:
                    data.work[-1].detail = company
                continue
            if "現在に至る" in str(company):
                continue
            data.work.append(WorkEntry(
                year=year, month=month, company=company,
                status=status or "", detail=detail
            ))

    wb.close()
    return data


def write_resume(filepath: str, data: ResumeData, output_path: str = None):
    """ResumeDataの内容でExcelを更新する"""
    if output_path and filepath != output_path:
        shutil.copy2(filepath, output_path)
        target = output_path
    else:
        target = filepath

    wb = load_workbook(target)
    ws = wb.active

    ws[CELL_MAP["furigana"]] = data.furigana
    ws[CELL_MAP["name"]] = data.name
    ws[CELL_MAP["gender"]] = data.gender
    if data.birthday:
        ws[CELL_MAP["birthday"]] = data.birthday
    ws[CELL_MAP["address"]] = data.address
    ws[CELL_MAP["nationality"]] = data.nationality
    ws[CELL_MAP["reg_number"]] = data.reg_number

    # 学歴の書き込み
    for i, edu in enumerate(data.education):
        row = EDUCATION_START_ROW + i
        if row >= WORK_HEADER_ROW:
            break
        ws.cell(row=row, column=1, value=edu.year)
        ws.cell(row=row, column=2, value=edu.month)
        ws.cell(row=row, column=3, value=edu.school)
        ws.cell(row=row, column=8, value=edu.status)

    # 職歴の書き込み
    row = WORK_START_ROW
    for work in data.work:
        ws.cell(row=row, column=1, value=work.year)
        ws.cell(row=row, column=2, value=work.month)
        ws.cell(row=row, column=3, value=work.company)
        ws.cell(row=row, column=8, value=work.status)
        row += 1
        if work.detail:
            ws.cell(row=row, column=3, value=work.detail)
            row += 1

    wb.save(target)
    wb.close()
    return target


def format_display_name(name: str, nationality: str, age: int) -> str:
    """表示用の名前をフォーマットする"""
    return f"{name}（{nationality}・{age}歳）"


def calculate_age(birthday: datetime) -> int:
    """生年月日から年齢を計算"""
    if not birthday:
        return 0
    today = datetime.today()
    age = today.year - birthday.year
    if (today.month, today.day) < (birthday.month, birthday.day):
        age -= 1
    return age
